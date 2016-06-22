#encoding=utf-8

import sys
import json
import os

from openpyxl.reader.excel import load_workbook


def readConfExecl(_fileName,outPath):
    wb = load_workbook(filename = _fileName )
    sheetnames = wb.sheetnames
    ws = wb.get_sheet_by_name(sheetnames[0])

    beginRow = 0
    beginColum = 0
    find = False
    for rx in range(1,ws.max_row + 1):
        for rc in range(1,ws.max_column + 1):
            value = ws.cell(row=rx,column=rc).value
            if value != None and value == "ID":
                find = True
                beginRow = rx
                beginColum = rc
                break
        if find is True:
            break

    key_value_dict = {}
    for rc in range(beginColum,ws.max_column + 1):
        value = ws.cell(row=beginRow,column=rc).value
        if value is not None :
            key_value_dict[rc] = value
    print key_value_dict
    main_dict = {}
    for rx in range(beginRow + 1,ws.max_row + 1):
        cell_value = {}
        ID = ws.cell(row=rx,column=beginColum).value
        for rc in range(beginColum,ws.max_column + 1):
            value = ws.cell(row=rx,column=rc).value
            if value is not None :
                cell_value[key_value_dict[rc]] = value

        main_dict[ID] = cell_value

    filename = os.path.basename(_fileName)
    name = filename[0:filename.index('.')]
    json.dump(main_dict,open(os.path.join(outpath,name + ".json"), 'w'),ensure_ascii = False)
    # print "readConfExecl"


if __name__ == '__main__':
    reload(sys)
    sys.setdefaultencoding( "utf-8" )
    if len(sys.argv) < 2:
        print "arguments error "
    else:
        outpath = "./"
        startPath = sys.argv[1]
        if len(sys.argv) >= 3:
            outpath = sys.argv[2]
        if not os.path.isabs(outpath):
            currentPath = startPath
            if not os.path.isdir(startPath):
                currentPath = os.path.dirname(startPath)
            outpath = os.path.join(currentPath,outpath)
        for (dirpath,dirnames,filenames) in os.walk(startPath):
            for filename in filenames:
                if filename.endswith('.xlsx'):
                    readConfExecl(os.path.join(dirpath,filename),outpath)
    # print "main"